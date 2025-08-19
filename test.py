import cv2
import face_recognition
import pickle
import os
import openpyxl
from datetime import datetime

# File paths
face_data_file = 'E:/Projects/attendance system/face_data.pkl'


attendance_file = 'attendance.xlsx'

# Load or initialize face data
face_data = pickle.load(open(face_data_file, 'rb')) if os.path.exists(face_data_file) else {'encodings': [], 'names': []}

# Initialize or create the Excel file for attendance
if not os.path.exists(attendance_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Date", "Time", "Status"])  # Header row
    wb.save(attendance_file)

def capture_face_and_name():
    """Capture a face and save it with a name."""
    cap = cv2.VideoCapture(0)
    print("Press 's' to save your face.")
    while True:
        _, img = cap.read()
        cv2.imshow("Capture Face", img)
        if cv2.waitKey(1) & 0xFF == ord('s'):
            img_small = cv2.cvtColor(cv2.resize(img, (0, 0), fx=0.25, fy=0.25), cv2.COLOR_BGR2RGB)
            encodings = face_recognition.face_encodings(img_small)
            if encodings:
                name = input("Enter name: ")
                face_data['encodings'].append(encodings[0])
                face_data['names'].append(name)
                with open(face_data_file, 'wb') as f:
                    pickle.dump(face_data, f)
                print(f"Saved face for {name}")
            else:
                print("No face detected. Try again.")
            break
    cap.release()
    cv2.destroyAllWindows()

def mark_attendance(name):
    """Mark the student's attendance in the Excel file."""
    attendance_file = 'E:/Projects/attendance system/attendance.xlsx'
    wb = openpyxl.load_workbook(attendance_file)
    ws = wb.active
    current_date = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")
    
    # Check if the student is already marked present
    already_present = any(row[0].value == name and row[1].value == current_date for row in ws.iter_rows(min_row=2))

    if not already_present:
        ws.append([name, current_date, current_time, "Present"])
        wb.save(attendance_file)
        print(f"Attendance marked for {name}")

def enhance_image(img):
    """Apply filters to enhance the image for better recognition accuracy."""
    # Convert to grayscale for noise reduction, then apply Gaussian blur
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    
    # Apply histogram equalization to enhance contrast
    equalized = cv2.equalizeHist(blurred)
    
    # Convert back to RGB
    enhanced_img = cv2.cvtColor(equalized, cv2.COLOR_GRAY2RGB)
    return enhanced_img

def recognize_face():
    """Identify faces in real-time from the webcam with enhanced image quality."""
    cap = cv2.VideoCapture(0)
    while True:
        _, img = cap.read()
        
        # Enhance image quality
        enhanced_img = enhance_image(img)
        
        # Resize and convert for face recognition
        img_small = cv2.cvtColor(cv2.resize(enhanced_img, (0, 0), fx=0.25, fy=0.25), cv2.COLOR_BGR2RGB)
        face_locs = face_recognition.face_locations(img_small)
        encodings = face_recognition.face_encodings(img_small, face_locs)
        
        for encoding, loc in zip(encodings, face_locs):
            matches = face_recognition.compare_faces(face_data['encodings'], encoding)
            distances = face_recognition.face_distance(face_data['encodings'], encoding)
            
            if True in matches:
                best_match_index = distances.argmin()
                confidence = (1 - distances[best_match_index]) * 100
                if confidence >= 45:
                    name = face_data['names'][best_match_index]
                    mark_attendance(name)  # Mark attendance in Excel
                else:
                    name = "Stranger"
                
                y1, x2, y2, x1 = [int(coord * 4) for coord in loc]
                color = (0, 255, 0) if name != "Stranger" else (0, 0, 255)
                cv2.rectangle(img, (x1, y1), (x2, y2), color, 2)
                cv2.putText(img, f"{name} ({round(confidence, 2)}%)", (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.75, color, 2)
            else:
                y1, x2, y2, x1 = [int(coord * 4) for coord in loc]
                cv2.rectangle(img, (x1, y1), (x2, y2), (0, 0, 255), 2)
                cv2.putText(img, "Stranger", (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 255), 2)
        
        cv2.imshow("Recognize Face", img)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    cap.release()
    cv2.destroyAllWindows()



# Usage
capture_face_and_name()  # Capture and save new face
recognize_face()         # Recognize faces from webcam