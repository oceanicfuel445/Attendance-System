import cv2
import numpy as np
import os
import pickle
from datetime import datetime
import openpyxl
from sklearn.svm import SVC
from sklearn.preprocessing import LabelEncoder

# File paths
face_data_file = "hog_svm_face_data.pkl"
attendance_file = "attendance.xlsx"

# Initialize or create the Excel file for attendance
if not os.path.exists(attendance_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Date", "Time", "Status"])  # Header row
    wb.save(attendance_file)

# HOG + SVM for face detection and recognition
def extract_hog_features(image):
    # Convert to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    # Initialize HOG descriptor
    hog = cv2.HOGDescriptor()
    # Compute HOG features
    hog_features = hog.compute(gray)
    return hog_features.flatten()

# Train the SVM model
def train_svm_model():
    if not os.path.exists(face_data_file):
        print("No face data found. Register faces first.")
        return

    with open(face_data_file, 'rb') as f:
        data = pickle.load(f)

    features = data['features']
    names = data['names']

    # Ensure all features have the same shape
    max_length = max(len(feature) for feature in features)
    padded_features = [np.pad(feature, (0, max_length - len(feature)), 'constant') for feature in features]

    features_array = np.array(padded_features)
    names_array = np.array(names)

    # Encode labels
    label_encoder = LabelEncoder()
    labels = label_encoder.fit_transform(names_array)

    # Train SVM classifier
    svm_model = SVC(kernel='linear', probability=True)
    svm_model.fit(features_array, labels)

    # Save the trained SVM model and label encoder
    with open('hog_svm_face_model.pkl', 'wb') as f:
        pickle.dump((svm_model, label_encoder), f)

    print("SVM model trained and saved.")

# Register a face with name
def capture_face_and_name():
    cap = cv2.VideoCapture(0)
    face_data = {'features': [], 'names': []}

    if os.path.exists(face_data_file):
        with open(face_data_file, 'rb') as f:
            face_data = pickle.load(f)

    print("Press 's' to save your face.")

    while True:
        _, img = cap.read()
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        for (x, y, w, h) in faces:
            face_img = img[y:y + h, x:x + w]
            hog_features = extract_hog_features(face_img)

            # Ensure all features have the same length
            if face_data['features']:
                max_length = max(len(feature) for feature in face_data['features'])
                hog_features = np.pad(hog_features, (0, max(0, max_length - len(hog_features))), 'constant')
            
            name = input("Enter your name: ")
            face_data['features'].append(hog_features)
            face_data['names'].append(name)

            with open(face_data_file, 'wb') as f:
                pickle.dump(face_data, f)

            print(f"Face data saved for {name}.")
            break

        cv2.imshow("Capture Face", img)
        if cv2.waitKey(1) & 0xFF == ord('s'):
            break

    cap.release()
    cv2.destroyAllWindows()

# Mark attendance in Excel
def mark_attendance(name):
    wb = openpyxl.load_workbook(attendance_file)
    ws = wb.active
    current_date = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")

    already_present = any(row[0].value == name and row[1].value == current_date for row in ws.iter_rows(min_row=2))

    if not already_present:
        ws.append([name, current_date, current_time, "Present"])
        wb.save(attendance_file)
        print(f"Attendance marked for {name}")
    else:
        print(f"{name} is already marked present today.")

# Recognize and mark attendance
def recognize_face():
    if not os.path.exists("hog_svm_face_model.pkl"):
        print("SVM model not found. Train the model first.")
        return

    with open("hog_svm_face_model.pkl", 'rb') as f:
        svm_model, label_encoder = pickle.load(f)

    cap = cv2.VideoCapture(0)
    print("Press 'q' to quit recognition.")

    while True:
        _, img = cap.read()
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        for (x, y, w, h) in faces:
            face_img = img[y:y + h, x:x + w]
            hog_features = extract_hog_features(face_img)

            # Ensure the feature vector has the same length as the training data
            if svm_model.n_features_in_ > len(hog_features):
                hog_features = np.pad(hog_features, (0, svm_model.n_features_in_ - len(hog_features)), 'constant')
            elif svm_model.n_features_in_ < len(hog_features):
                hog_features = hog_features[:svm_model.n_features_in_]

            prediction = svm_model.predict([hog_features])
            name = label_encoder.inverse_transform(prediction)[0]
            mark_attendance(name)

            color = (0, 255, 0) if name != "Stranger" else (0, 0, 255)
            cv2.rectangle(img, (x, y), (x + w, y + h), color, 2)
            cv2.putText(img, name, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.75, color, 2)

        cv2.imshow("Recognize Face", img)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

# Main menu
while True:
    print("\nChoose an option:")
    print("1. Register a new face.")
    print("2. Train SVM model.")
    print("3. Recognize and mark attendance.")
    print("4. Exit")
    choice = input("Enter your choice: ")

    if choice == '1':
        capture_face_and_name()
    elif choice == '2':
        train_svm_model()
    elif choice == '3':
        recognize_face()
    elif choice == '4':
        print("Exiting the program.")
        break
    else:
        print("Invalid choice. Please try again.")